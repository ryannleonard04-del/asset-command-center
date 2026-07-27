import os
from html import escape

from flask import Flask, request, redirect, session, url_for

import psycopg2
import psycopg2.extras

import csv
import io

from openpyxl import load_workbook
import xlrd

from flask_login import LoginManager, login_user, logout_user, login_required
from flask_login import UserMixin, current_user
from werkzeug.security import generate_password_hash, check_password_hash

# App setup and session security
app = Flask(
    __name__,
    static_folder="Image",
    static_url_path="/Image"
)

app.secret_key = os.environ.get("ITAM_SECRET_KEY")

if not app.secret_key:
    raise RuntimeError("ITAM_SECRET_KEY environment variable must be set")

app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "True").lower() in ("true", "1", "yes")
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

class User(UserMixin):
    def __init__(self, user_id, email, username, role):
        self.id = str(user_id)
        self.email = email
        self.username = username
        self.role = role


@login_manager.user_loader
def load_user(user_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        SELECT user_id, email, username, role
        FROM public.app_users
        WHERE user_id = %s
          AND is_active = true
    """, (user_id,))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if row:
        return User(row["user_id"], row["email"], row["username"], row["role"])

    return None
DB_HOST = os.environ.get("ITAM_DB_HOST", "localhost")
DB_PORT = os.environ.get("ITAM_DB_PORT", "5432")
DB_NAME = os.environ.get("ITAM_DB_NAME", "ITAM")
DB_USER = os.environ.get("ITAM_DB_USER", "postgres")
DB_PASSWORD = os.environ.get("ITAM_DB_PASSWORD")


def get_db_connection():
    # Create a new database connection for each request.
    # This avoids sharing state between requests and keeps queries isolated.
    return psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD
    )

def page(title, body):
    return f"""
    <html>
    <head>
        <title>{title}</title>
        <style>
            :root {{
                --background: #06111f;
                --surface: #0b1f33;
                --surface-alt: #102a43;
                --border: #1e4f6b;
                --text: #f8fafc;
                --muted: #a8c0d6;
                --accent: #38bdf8;
                --action: #0284c7;
                --danger: #ef4444;
                --warning: #f59e0b;
                --success: #22c55e;
                --info: #14b8a6;
            }}

            body {{
                font-family: Segoe UI, Arial, sans-serif;
                background: var(--background);
                color: var(--text);
                margin: 0;
                font-size: 12pt;
            }}

            .header {{
    background: linear-gradient(135deg, #071827, #0b4662);
    color: white;
    padding: 35px;
    border-bottom: 1px solid rgba(56, 189, 248, 0.35);
}}

.header h1 {{
    margin: 0 0 18px;
    line-height: 1.2;
}}

.header a {{
    color: #bae6fd;
}}

.content {{
    padding: 30px;
}}
            .grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
                gap: 22px;
            }}

            .card {{
                background: var(--surface);
                padding: 25px;
                border: 1px solid rgba(56, 189, 248, 0.22);
                border-radius: 16px;
                box-shadow: 0 10px 28px rgba(0, 0, 0, 0.25);
            }}

            .btn,
            button {{
                background: var(--action);
                color: white;
                padding: 11px 18px;
                text-decoration: none;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                cursor: pointer;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                min-height: 44px;
                font-size: 12pt;
                transition: background 0.2s ease, transform 0.2s ease;
            }}

            .btn:hover,
            button:hover {{
                background: var(--accent);
                color: #06111f;
                transform: translateY(-1px);
            }}

            .btn.small {{
                padding: 6px 10px;
                font-size: 12pt;
                border-radius: 6px;
                min-width: 90px;
                height: 36px;
            }}

            .toolbar {{
                display: flex;
                flex-wrap: wrap;
                gap: 12px;
                align-items: center;
                justify-content: space-between;
                margin-bottom: 18px;
            }}

            .toolbar .search-group {{
                display: flex;
                gap: 10px;
                align-items: center;
                flex: 1;
                min-width: 0;
            }}

            .toolbar .search-group form {{
                display: flex;
                gap: 10px;
                align-items: center;
                flex: 1;
                min-width: 0;
            }}

            .toolbar .search-group input[type="text"] {{
                flex: 1;
                min-width: 220px;
                max-width: 360px;
                height: 44px;
            }}

            .toolbar .actions-group {{
                display: flex;
                gap: 10px;
                align-items: center;
                justify-content: flex-end;
                flex-shrink: 0;
            }}

            .toolbar .actions-group .btn {{
                white-space: nowrap;
            }}

            .toolbar .search-group button,
.toolbar .search-group .btn,
.toolbar .actions-group .btn {{
    width: 130px;
    min-width: 130px;
    max-width: 130px;
    height: 44px;
    min-height: 44px;
    padding: 0 12px;
    box-sizing: border-box;
    text-align: center;
    font-size: 14px;
    white-space: nowrap;
}}

            th.actions,
            td.actions {{
                text-align: center;
                width: 140px;
            }}

            input,
            select,
            textarea {{
                padding: 10px;
                width: 300px;
                border-radius: 8px;
                border: 1px solid #0ea5e9;
                background: #071525;
                color: white;
                box-sizing: border-box;
            }}

            input:focus,
            select:focus,
            textarea:focus {{
                outline: none;
                border-color: #67e8f9;
                box-shadow: 0 0 0 3px rgba(34, 211, 238, 0.16);
            }}

            .back {{
                display: inline-block;
                margin-bottom: 20px;
                text-decoration: none;
                font-weight: bold;
                color: var(--accent);
            }}

            .scroll {{
                overflow: auto;
                max-height: 75vh;
                background: #f8fafc;
                border: 1px solid var(--border);
                border-radius: 8px;
            }}

            .dashboard-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
                gap: 18px;
                margin-bottom: 24px;
            }}

            .stat-card {{
                background: var(--surface);
                border: 1px solid rgba(56, 189, 248, 0.25);
                border-radius: 18px;
                padding: 24px;
                box-shadow: 0 14px 35px rgba(0, 0, 0, 0.25);
            }}

            .stat-card h3 {{
                margin: 0 0 10px;
                font-size: 16px;
                color: #dbeafe;
            }}

            .stat-card .value {{
                font-size: 36px;
                font-weight: 800;
                color: var(--accent);
            }}

            .stat-card .label {{
                font-size: 13px;
                color: var(--muted);
                margin-top: 8px;
            }}

            .dashboard-panel {{
                display: grid;
                grid-template-columns: 2fr 1fr;
                gap: 20px;
                align-items: start;
            }}

            .chart-card {{
                background: var(--surface);
                border: 1px solid rgba(56, 189, 248, 0.25);
                border-radius: 18px;
                padding: 22px;
                box-shadow: 0 12px 30px rgba(0, 0, 0, 0.22);
            }}

            .chart-card h2 {{
                margin-top: 0;
                color: #dbeafe;
            }}

            .chart-list {{
                list-style: none;
                padding: 0;
                margin: 0;
            }}

            .chart-item {{
                display: grid;
                grid-template-columns: 1fr auto;
                gap: 12px;
                align-items: center;
                margin-bottom: 14px;
            }}

            .chart-label {{
                font-size: 14px;
                color: #dbeafe;
            }}

            .chart-value {{
                font-size: 14px;
                color: var(--accent);
                font-weight: 700;
            }}

            .chart-bar-container {{
                background: #162b3d;
                border-radius: 999px;
                height: 12px;
                overflow: hidden;
            }}

            .chart-bar {{
                height: 100%;
                border-radius: 999px;
                background: linear-gradient(90deg, #0ea5e9, #67e8f9);
            }}

            .quick-links {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
                gap: 14px;
                margin-top: 24px;
            }}

            .quick-link {{
                background: var(--surface);
                border: 1px solid rgba(56, 189, 248, 0.25);
                border-radius: 16px;
                padding: 18px;
                text-decoration: none;
                color: white;
                display: flex;
                flex-direction: column;
                gap: 10px;
                transition: background 0.2s ease, transform 0.2s ease;
            }}

            .quick-link:hover {{
                background: var(--surface-alt);
                border-color: var(--accent);
                transform: translateY(-2px);
            }}

            .quick-link h3 {{
                margin: 0;
                font-size: 16px;
                color: #dbeafe;
            }}

            .quick-link p {{
                margin: 0;
                color: var(--muted);
            }}

            table {{
                width: 100%;
                border-collapse: collapse;
                background: #f8fafc;
                color: #0f172a;
                font-size: 13px;
            }}

            th {{
                background: #075985;
                color: white;
                padding: 10px;
                text-align: left;
                position: sticky;
                top: 0;
                z-index: 2;
                white-space: nowrap;
            }}

            td {{
                padding: 8px;
                border-bottom: 1px solid #cbd5e1;
                white-space: nowrap;
                color: #0f172a;
            }}

            tbody tr:nth-child(odd) td {{
                background: #f8fafc;
                color: #0f172a;
            }}

            tbody tr:nth-child(even) td {{
                background: #eaf2f8;
                color: #0f172a;
            }}

            tbody tr:hover td {{
                background: #dbeafe;
                color: #0f172a;
            }}

            @media (max-width: 850px) {{
                .dashboard-panel {{
                    grid-template-columns: 1fr;
                }}

                .content {{
                    padding: 18px;
                }}
            }}
        </style>
    </head>
    <body>{body}</body>
    </html>
    """

def require_admin():
    # Restrict certain routes to Admin users only.
    # Returns an access-denied page if the current user lacks privileges.
    if current_user.role != "Admin":
        return page(
            "Access Denied",
            """
            <div class="header">
                <h1>Access Denied</h1>
            </div>
            <div class="content">
                <p>Admin access is required for this action.</p>
                <a class="back" href="/">← Dashboard</a>
            </div>
            """
        )

    return None


def read_upload_rows(uploaded_file):
    # Normalize Excel and CSV uploads into the same header/row structure.
    # Supports .xlsx, .xls, and .csv, returning parsed headers and rows.
    filename = (uploaded_file.filename or "").lower()

    if filename.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            header = cell.value
            if header is None:
                continue

            header = str(header).strip()
            if header:
                headers.append(header)

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for index, header in enumerate(headers):
                if index < len(row):
                    row_data[header] = row[index]
            rows.append(row_data)

        return headers, rows, None

    if filename.endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=uploaded_file.read())
        sheet = workbook.sheet_by_index(0)

        headers = []
        for index in range(sheet.ncols):
            header = str(sheet.cell_value(0, index)).strip()
            if header:
                headers.append(header)

        rows = []
        for row_index in range(1, sheet.nrows):
            row_data = {}
            for col_index, header in enumerate(headers):
                if col_index < sheet.ncols:
                    row_data[header] = sheet.cell_value(row_index, col_index)
            rows.append(row_data)

        return headers, rows, None

    if filename.endswith(".csv"):
        raw_text = uploaded_file.read().decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(raw_text))

        headers = []
        for header in reader.fieldnames or []:
            if header is None:
                continue

            header = header.strip()
            if header:
                headers.append(header)

        rows = []
        for raw_row in reader:
            row_data = {}
            for key, value in raw_row.items():
                if key is None:
                    continue

                normalized_key = key.strip()
                if normalized_key:
                    row_data[normalized_key] = value

            rows.append(row_data)

        return headers, rows, None

    return None, None, "Unsupported file type. Please upload .xlsx, .xlsm, or .csv."


def is_allowed_upload_file(uploaded_file):
    # Reject uploads that do not match the supported spreadsheet formats.
    filename = (uploaded_file.filename or "").lower()
    # Limit uploads to CSV only as requested.
    return filename.endswith(".csv")


def hash_password(raw_password):
    # Hash passwords securely using PBKDF2.
    # This should be used when creating or updating account passwords.
    return generate_password_hash(raw_password, method="pbkdf2:sha256", salt_length=16)


def password_matches(stored_password, entered_password):
    # Accept hashed passwords first, then fall back to legacy plain text for older accounts.
    if not stored_password:
        return False

    if stored_password.startswith(("pbkdf2:", "scrypt:", "bcrypt:")):
        return check_password_hash(stored_password, entered_password)

    return stored_password == entered_password


def render_edit_page(title, icon, back_href, heading, record, editable_fields):
    # Reuse one simple form renderer for all table-specific edit pages.
    # This helper builds a consistent edit form for staff, students, and hotspots.
    from html import escape

    form_fields = ""
    for field in editable_fields:
        field_value = record.get(field, "")
        if field_value is None:
            field_value = ""

        form_fields += f"""
        <p><b>{escape(field)}</b></p>
        <input type="text" name="{escape(field)}" value="{escape(str(field_value), quote=True)}" style="width:400px;padding:8px;">
        <br><br>
        """

    return page(
        title,
        f"""
        <div class="header">
            <h1>{icon} {escape(title)}</h1>
        </div>

        <div class="content">
            <a class="back" href="{back_href}">← Back</a>
            <h2>{escape(str(heading))}</h2>

            <form method="POST">
                {form_fields}
                <button type="submit">Save Changes</button>
            </form>
        </div>
        """
    )


def save_edit_fields(table_name, pk_name, pk_value, editable_fields):
    # Update only the requested row with the submitted values.
    # This helper is used by edit routes with a shared update pattern.
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        set_clause = ", ".join([f'"{field}" = %s' for field in editable_fields])
        values = [request.form.get(field, "").strip() for field in editable_fields]
        values.append(pk_value)

        cur.execute(
            f'UPDATE public."{table_name}" SET {set_clause} WHERE "{pk_name}" = %s',
            values
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()


def fetch_single_record(table_name, pk_name, pk_value, columns=None):
    # Fetch the current row so edit pages can be pre-filled.
    # Returns a dictionary for the requested table and primary key.
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        selected_columns = "*" if not columns else ", ".join([f'"{column}"' for column in columns])
        cur.execute(
            f'SELECT {selected_columns} FROM public."{table_name}" WHERE "{pk_name}" = %s',
            (pk_value,)
        )
        return cur.fetchone()
    finally:
        cur.close()
        conn.close()

def make_table_page(title, icon, table_name, columns):
    # Render a generic table view for the supported inventory tables.
    # This function handles search, row rendering, and header actions.
    from html import escape

    allowed_tables = {
        "devices",
        "devices_smart",
        "movements",
        "staff",
        "students",
        "hotspots",
        "depreciation",
        "repairs"
    }

    if table_name not in allowed_tables:
        return page(
            "Error",
            """
            <div class="header">
                <h1>Table Error</h1>
            </div>
            <div class="content">
                <p>The requested table is not allowed.</p>
                <a class="back" href="/">← Dashboard</a>
            </div>
            """
        )

    search = request.args.get("search", "").strip()

    # Actions is displayed on the page but is not a database column.
    database_columns = [
        column for column in columns
        if column != "Actions"
    ]

    selected_columns = ", ".join(
        f'"{column}"' for column in database_columns
    )

    conn = get_db_connection()

    cur = conn.cursor(
        cursor_factory=psycopg2.extras.RealDictCursor
    )

    try:
        if search:
            search_conditions = " OR ".join(
                f'CAST("{column}" AS TEXT) ILIKE %s'
                for column in database_columns
            )

            search_values = [
                f"%{search}%"
                for column in database_columns
            ]

            query = f"""
                SELECT {selected_columns}
                FROM public."{table_name}"
                WHERE {search_conditions}
            """

            cur.execute(query, search_values)

        else:
            query = f"""
                SELECT {selected_columns}
                FROM public."{table_name}"
            """

            cur.execute(query)

        rows = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    header_cells = ""

    for column in columns:
        header_cells += f"<th>{escape(column)}</th>"

    table_rows = ""
    row_index = 0

    for row in rows:
        row_index += 1
        table_rows += "<tr>"

        for column in columns:
            # Show a user-friendly sequential ID in the grid for devices
            if table_name == "devices" and column == "DeviceID":
                displayed_value = str(row_index)
            else:
                value = row.get(column)

                if value is None:
                    displayed_value = ""
                elif hasattr(value, "strftime"):
                    try:
                        displayed_value = value.strftime(
                            "%m/%d/%Y"
                        )
                    except Exception:
                        displayed_value = str(value)
                else:
                    displayed_value = str(value)

            table_rows += (
                f"<td>{escape(displayed_value)}</td>"
            )

        table_rows += "</tr>"

    if not table_rows:
        table_rows = f"""
            <tr>
                <td colspan="{len(columns)}">
                    No records found.
                </td>
            </tr>
        """

    safe_title = escape(title)
    safe_search = escape(search, quote=True)
    hotspot_edit_button = ""

    action_lookup_pages = {
    "devices": ("/devices/edit", "Add/Edit Device"),
    "staff": ("/staff/edit", "Add/Edit Staff"),
    "students": ("/students/edit", "Add/Edit Students"),
    "hotspots": ("/hotspots/edit", "Add/Edit Hotspots"),
    "repairs": ("/repairs/add", "Add Repair"),
}

    add_buttons = ""
    if table_name in action_lookup_pages:
        lookup_path, label = action_lookup_pages[table_name]
        add_buttons = f"""
            <a class="btn" href="{lookup_path}" style="margin-left:10px;">
                {label}
            </a>
        """

    body = f"""
        <div class="header">
            <h1>{icon} {safe_title}</h1>
        </div>

        <div class="content">

            <a class="back" href="/">
                ← Dashboard
            </a>

            <div class="toolbar">
                <div class="search-group">
                    <form method="GET">
                        <input
                            type="text"
                            name="search"
                            value="{safe_search}"
                            placeholder="Search records"
                        >

                        <button type="submit">
                            Search
                        </button>

                        <a
                            class="btn"
                            href="{request.path}"
                        >
                            Clear
                        </a>
                    </form>
                </div>

                <div class="actions-group">
                    {add_buttons}
                </div>
            </div>

            <p>
                <b>{len(rows)}</b> records displayed
            </p>

            <div class="scroll">
                <table>
                    <thead>
                        <tr>
                            {header_cells}
                        </tr>
                    </thead>

                    <tbody>
                        {table_rows}
                    </tbody>
                </table>
            </div>

        </div>
    """

    return page(title, body)

@app.route("/login", methods=["GET", "POST"])
def login():
    # Authenticate the user and create a session when credentials match.
    # The login page is public and starts a Flask-Login session.
    error = ""

    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()

        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cur.execute("""
            SELECT user_id, email, username, role, password_hash
            FROM public.app_users
            WHERE email = %s
              AND is_active = true
        """, (email,))

        row = cur.fetchone()
        cur.close()
        conn.close()

        if row and password_matches(row["password_hash"], password):
            user = User(row["user_id"], row["email"], row["username"], row["role"])
            login_user(user)
            return redirect("/")
        else:
            error = "Invalid email or password."

    return f"""
    <!-- Login page uses the branded background and centered auth card. -->
    <html>
    <head>
        <title>ITAM Login</title>
        <style>
            body {{
                margin: 0;
                min-height: 100vh;
                background-image: url('/Image/asset-command-center-login.png');
                background-size: cover;
                background-position: center center;
                background-repeat: no-repeat;
                font-family: Segoe UI, Arial, sans-serif;
                display: flex;
                align-items: center;
                justify-content: center;
            }}

            .box {{
                background: transparent;
                color: white;
                padding: 35px;
                border-radius: 0;
                border: none;
                box-shadow: none;
                 width: 350px;
            }}

            h1 {{
    text-align: center;
    color: #38bdf8;
    margin-top: 0;
}}

.subtitle {{
    text-align: center;
    margin-bottom: 20px;
    color: #dbeafe;
}}

input {{
    width: 100%;
    padding: 12px;
    margin: 10px 0;
    border: 1px solid #0ea5e9;
    border-radius: 8px;
    background: #071525;
    color: white;
    box-sizing: border-box;
}}

button {{
    width: 100%;
    padding: 12px;
    background: #0284c7;
    color: white;
    border: none;
    border-radius: 8px;
    font-weight: bold;
    margin-top: 10px;
    cursor: pointer;
}}

button:hover {{
    background: #0ea5e9;
}}

.error {{
    color: #f87171;
    text-align: center;
    font-weight: bold;
}}
</style>
</head>
    <body>
    <div class="box">
        <h1>ASSET COMMAND CENTER</h1>
        <div class="subtitle">TRACK EVERY ASSET. FOLLOW EVERY MOVEMENT.</div>

        <p class="error">{error}</p>

        <form method="POST">
            <input type="email" name="email" placeholder="Email Address" required>
            <input type="password" name="password" placeholder="Password" required>
            <button type="submit">🔐 SIGN IN</button>
        </form>
    </div>
</body>
</html>
    """
@app.route("/logout")
@login_required
def logout():
    # Log out the current user and return to the login page.
    logout_user()
    return redirect("/login")

@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():

    message = ""

    if request.method == "POST":

        new_password = request.form.get("new_password", "").strip()

        if new_password:
            conn = get_db_connection()
            cur = conn.cursor()

            cur.execute(
                """
                UPDATE public.app_users
                SET password_hash = %s
                WHERE user_id = %s
                """,
                (hash_password(new_password), current_user.id)
            )

            conn.commit()
            cur.close()
            conn.close()

            message = "Password updated successfully."
        else:
            message = "Password update failed. Please enter a new password."

    return f"""
    <html>
    <head>
        <title>Change Password</title>
    </head>
    <body style="font-family:Segoe UI;padding:40px;">
        <h1>Change Password</h1>

        <p style="color:green;">{message}</p>

        <form method="POST">
            <input
                type="password"
                name="new_password"
                placeholder="New Password"
                required
                style="padding:10px;width:300px;"
            >

            <br><br>

            <button type="submit">
                Update Password
            </button>
        </form>

        <br><br>

        <a href="/">Back to Dashboard</a>
    </body>
    </html>
    """
@app.route("/admin/users/add", methods=["GET", "POST"])
@login_required
def add_user():
    # Add a new application user. Admin access is required.

    if current_user.role != "Admin":
        return "Access denied. Admins only."

    message = ""

    if request.method == "POST":

        username = request.form.get("username", "")
        email = request.form.get("email", "")
        password = request.form.get("password", "")
        role = request.form.get("role", "Editor")

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            """
            INSERT INTO public.app_users
            (
                username,
                email,
                password_hash,
                role,
                is_active
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                true
            )
            """,
            (
                username,
                email,
                hash_password(password),
                role
            )
        )

        conn.commit()

        cur.close()
        conn.close()

        message = "User created successfully."

    return f"""
    <div class="header">
        <h1>➕ Add User</h1>
    </div>

    <div class="content">

        <a class="back" href="/admin/users">← User Management</a>

        <p style="color:green;">{message}</p>

        <form method="POST">

            <p>Name</p>
            <input type="text" name="username" required>

            <p>Email</p>
            <input type="email" name="email" required>

            <p>Temporary Password</p>
            <input type="text" name="password" required>

            <p>Role</p>
            <select name="role">
                <option value="Editor">Editor</option>
                <option value="Admin">Admin</option>
            </select>

            <br><br>

            <button type="submit">
                Create User
            </button>

        </form>

    </div>
    """
@app.route("/admin/users/reset/<int:user_id>")
@login_required
def reset_user_password(user_id):
    # Reset a user's password back to a default value.

    if current_user.role != "Admin":
        return "Access denied."

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        """
        UPDATE public.app_users
        SET password_hash = %s
        WHERE user_id = %s
        """,
        (hash_password('ChangeMe123'), user_id)
    )

    conn.commit()

    cur.close()
    conn.close()

    return redirect("/admin/users?msg=password_reset")

@app.route("/admin/users/toggle/<int:user_id>")
@login_required
def toggle_user(user_id):
    # Enable or disable a user account. Prevents the current admin from disabling themselves.

    if current_user.role != "Admin":
        return "Access denied."

    if int(current_user.id) == user_id:
        return "You cannot disable your own account."

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute(
        """
        SELECT is_active
        FROM public.app_users
        WHERE user_id = %s
        """,
        (user_id,)
    )

    row = cur.fetchone()

    new_status = not row["is_active"]

    cur.execute(
        """
        UPDATE public.app_users
        SET is_active = %s
        WHERE user_id = %s
        """,
        (new_status, user_id)
    )

    conn.commit()

    cur.close()
    conn.close()

    return redirect("/admin/users")

@app.route("/admin/users/edit/<int:user_id>", methods=["GET", "POST"])
@login_required
def edit_user(user_id):
    # Edit an existing application user and optionally update the password.
    # This route is only available to Admin users.

    if current_user.role != "Admin":
        return "Access denied."

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        username = request.form.get("username", "")
        email = request.form.get("email", "")
        role = request.form.get("role", "Editor")
        password = request.form.get("password", "")

        if password:
            cur.execute("""
                UPDATE public.app_users
                SET username = %s, email = %s, role = %s, password_hash = %s
                WHERE user_id = %s
            """, (username, email, role, hash_password(password), user_id))
        else:
            cur.execute("""
                UPDATE public.app_users
                SET username = %s, email = %s, role = %s
                WHERE user_id = %s
            """, (username, email, role, user_id))

        conn.commit()
        cur.close()
        conn.close()

        return redirect("/admin/users")

    cur.execute("""
        SELECT user_id, username, email, role, is_active
        FROM public.app_users
        WHERE user_id = %s
    """, (user_id,))

    user = cur.fetchone()

    cur.close()
    conn.close()

    return f"""
    <div class="header">
        <h1>✏️ Edit User</h1>
    </div>

    <div class="content">
        <a class="back" href="/admin/users">← User Management</a>

        <form method="POST">
            <p>Name</p>
            <input type="text" name="username" value="{user['username']}" required>

            <p>Email</p>
            <input type="email" name="email" value="{user['email']}" required>

            <p>Role</p>
            <select name="role">
                <option value="Editor" {"selected" if user["role"] == "Editor" else ""}>Editor</option>
                <option value="Admin" {"selected" if user["role"] == "Admin" else ""}>Admin</option>
            </select>

            <p>New Password</p>
            <input type="text" name="password" placeholder="Leave blank to keep current password">

            <br><br>

            <button type="submit">Save Changes</button>
        </form>
    </div>
    """

@app.route("/admin/users")
@login_required
def admin_users():
    # Display the admin user management panel with edit, reset, and toggle actions.
    # Only Admin users can view and manage application user accounts.

    message = ""

    if request.args.get("msg") == "password_reset":
        message = "✅ Password reset successfully."

    if current_user.role != "Admin":
        return "Access denied. Admins only."

  
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        SELECT user_id, username, email, role, is_active
        FROM public.app_users
        ORDER BY username
    """)

    rows = cur.fetchall()
    cur.close()
    conn.close()

    table_rows = ""

    for row in rows:
       table_rows += f"""
<tr>
    <td>{escape(str(row.get('user_id')))}</td>
    <td>{escape(str(row.get('username') or ''))}</td>
    <td>{escape(str(row.get('email') or ''))}</td>
    <td>{escape(str(row.get('role') or ''))}</td>
    <td>{escape(str(row.get('is_active')))}</td>

    <td class=\"actions\">
        <a class=\"btn small\" href=\"/admin/users/edit/{escape(str(row.get('user_id')))}\">Edit</a>
        <a class=\"btn small\" href=\"/admin/users/reset/{escape(str(row.get('user_id')))}\">Reset</a>
        <a class=\"btn small\" href=\"/admin/users/toggle/{escape(str(row.get('user_id')))}\">Enable/Disable</a>
    </td>

</tr>
"""

    body = f"""
    <div class="header">
        <h1>👤 User Management</h1>
        <p>Admins only</p>
    </div>

    <div class="content">
        <a class="back" href="/">← Dashboard</a>

        <h3>System Users</h3>
        <p style="color:green;font-weight:bold;">
    {escape(message)}
                <th>Role</th>
                <th>Active</th>
<th>Actions</th>
            </tr>
            {table_rows}
        </table>
    </div>
    """

    return page("User Management", body)

@app.route("/")
@login_required
def home():
    # Dashboard landing page that shows system metrics, status breakdowns, and quick action cards.
    # This section queries the database for live summary totals used by the dashboard widgets.
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute('SELECT COUNT(*) AS total FROM public.devices')
    device_count = cur.fetchone()["total"]

    cur.execute('SELECT COUNT(*) AS total FROM public.movements')
    movement_count = cur.fetchone()["total"]

    cur.execute('SELECT COUNT(*) AS total FROM public.staff')
    staff_count = cur.fetchone()["total"]

    cur.execute('SELECT COUNT(*) AS total FROM public.students')
    student_count = cur.fetchone()["total"]

    cur.execute('SELECT COUNT(*) AS total FROM public.hotspots')
    hotspot_count = cur.fetchone()["total"]

    cur.execute('SELECT "Status", COUNT(*) AS count FROM public.devices GROUP BY "Status" ORDER BY count DESC')
    status_rows = cur.fetchall()

    cur.close()
    conn.close()

    status_rows = status_rows or []
    total_status = sum([row.get("count", 0) for row in status_rows]) or 1

    status_items = ""
    for row in status_rows:
        status = row.get("Status") or "Unknown"
        count = row.get("count", 0)
        width = int((count / total_status) * 100)
        status_items += f"""
            <li class=\"chart-item\">
                <span class=\"chart-label\">{status}</span>
                <span class=\"chart-value\">{count}</span>
                <div class=\"chart-bar-container\"><div class=\"chart-bar\" style=\"width:{width}%\"></div></div>
            </li>
        """

    body = f"""
<div class="header">
    <h1>IT Asset Management Dashboard</h1>
    <p>Welcome, {escape(str(current_user.username))} ({escape(str(current_user.role))})</p>
    <p>Live asset, assignment, and movement insights across the ITAM system.</p>
    <p>
        <a style="color:white;font-weight:bold;" href="/logout">Logout</a>
        |
        <a style="color:white;font-weight:bold;" href="/change-password">Change Password</a>
    </p>
</div>

<div class="content">
    <div class="dashboard-grid">
        <div class="stat-card">
            <h3>Total Devices</h3>
            <div class="value">{device_count}</div>
            <div class="label">All inventory records in the system</div>
        </div>
        <div class="stat-card">
            <h3>Total Movements</h3>
            <div class="value">{movement_count}</div>
            <div class="label">Assignment and transfer transactions</div>
        </div>
        <div class="stat-card">
            <h3>Total Staff</h3>
            <div class="value">{staff_count}</div>
            <div class="label">Active staff members in the directory</div>
        </div>
        <div class="stat-card">
            <h3>Total Students</h3>
            <div class="value">{student_count}</div>
            <div class="label">Student records available to assign</div>
        </div>
    </div>

    <div class="dashboard-panel">
        <div class="chart-card">
            <h2>Device Status Breakdown</h2>
            <ul class="chart-list">
                {status_items}
            </ul>
        </div>
        <div class="chart-card">
            <h2>Quick Summary</h2>
            <div class="stat-card" style="margin:0;padding:18px;background:#161616;">
                <h3>Hotspot Records</h3>
                <div class="value">{hotspot_count}</div>
                <div class="label">Hotspot movement entries</div>
            </div>
            <div class="stat-card" style="margin-top:16px;padding:18px;background:#161616;">
                <h3>Actions</h3>
                <div class="label">Jump to key workflows</div>
                <div style="display:flex;flex-direction:column;gap:10px;margin-top:14px;">
                    <a class="btn" href="/devices">Devices</a>
                    <a class="btn" href="/movement-entry">Movement Entry</a>
                    <a class="btn" href="/bulk-upload">Bulk Upload</a>
                </div>
            </div>
        </div>
    </div>

    <div class="quick-links">
        <a class="quick-link" href="/devices"><h3>Device Inventory</h3><p>View and search all devices.</p></a>
        <a class="quick-link" href="/movements"><h3>Movement History</h3><p>Review recent transactions.</p></a>
        <a class="quick-link" href="/staff"><h3>Staff Directory</h3><p>Manage staff records.</p></a>
        <a class="quick-link" href="/students"><h3>Student Directory</h3><p>Manage student records.</p></a>
        <a class="quick-link" href="/hotspots"><h3>Hotspot Inventory</h3><p>Track wireless device movements.</p></a>
        <a class="quick-link" href="/repairs"><h3>Repair Tracking</h3><p>Track repair tickets, vendors, costs, and resolution details.</p></a>
        <a class="quick-link" href="/admin/users"><h3>User Management</h3><p>Admin controls for app accounts.</p></a>
    </div>
</div>
"""

    return page("ITAM Asset Manager", body)

@app.route("/devices/edit/<int:device_id>", methods=["GET", "POST"])
@login_required
def edit_device(device_id):
    # Edit an existing device record by DeviceID.
    # Handles both GET form rendering and POST saves.

    editable_fields = [
        "AssetTag", "DeviceType", "Manufacturer", "Model",
        "PurchaseDate", "WarrantyEnd", "AssignedTo", "AssignmentDate",
        "Location", "Condition", "ReplacementEligible", "NumberofReplacement",
        "AssignedFlag", "WarrantyStatus", "WarrantyCovered",
        "WarrantyDaysRemaining", "WarrantyExpiring90Days", "ActionRequired",
        "Cost", "Notes", "IsActive", "UserType", "Status"
    ]

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        for field in editable_fields:
            cur.execute(
                f'UPDATE public.devices SET "{field}" = %s WHERE "DeviceID" = %s',
                (request.form.get(field, ""), device_id)
            )

        conn.commit()
        cur.close()
        conn.close()

        return redirect("/devices")

    cur.execute(
        """
        SELECT *
        FROM public.devices
        WHERE "DeviceID" = %s
        """,
        (device_id,)
    )

    device = cur.fetchone()

    cur.close()
    conn.close()

    if not device:
        return "Device not found."

    form_fields = ""

    for field in editable_fields:
        form_fields += f"""
        <p><b>{field}</b></p>
        <input type="text" name="{field}" value="{device.get(field, '')}" style="width:400px;padding:8px;">
        <br><br>
        """

    return f"""
    <div class="header">
        <h1>✏️ Edit Device</h1>
    </div>

    <div class="content">
        <a class="back" href="/devices">← Devices</a>

        <h2>{device.get('AssetTag')}</h2>

        <form method="POST">
            {form_fields}

            <button type="submit">Save Changes</button>
        </form>
    </div>
    """


@app.route("/devices/edit", methods=["GET", "POST"])
@login_required
def edit_device_lookup():
    # Lookup page for device edit by Device ID.
    # Allows users to open an existing device or navigate to the add-new-device form.
    if request.method == "POST":
        device_id = request.form.get("device_id", "").strip()
        if not device_id:
            return page(
                "Edit Device",
                """
                <div class="header"><h1>💻 Add/Edit Device</h1></div>
                <div class="content">
                    <a class="back" href="/devices">← Devices</a>
                    <p>Please enter a Device ID to edit.</p>
                    <form method="POST">
                        <p><b>Device ID</b></p>
                        <input type="text" name="device_id" style="width:400px;padding:8px;">
                        <br><br>
                        <button type="submit">Open Device</button>
                    </form>
                    <br>
                    <a class="btn" href="/devices/add">Add New Device</a>
                </div>
                """
            )

        return redirect(f"/devices/edit/{device_id}")

    return page(
        "Add/Edit Device",
        """
        <div class="header"><h1>💻 Add/Edit Device</h1></div>
        <div class="content">
            <a class="back" href="/devices">← Devices</a>
            <p>Enter a Device ID to edit an existing device, or add a new one.</p>
            <form method="POST" style="margin-bottom:20px;">
                <p><b>Device ID</b></p>
                <input type="text" name="device_id" style="width:400px;padding:8px;">
                <br><br>
                <button type="submit">Open Device</button>
            </form>
            <a class="btn" href="/devices/add">Add New Device</a>
        </div>
        """
    )


@app.route("/devices/add", methods=["GET", "POST"])
@login_required
def add_device():
    # Provide a complete form to add a new device inventory record.
    # Validates that DeviceID is present before inserting the row.
    editable_fields = [
        "DeviceID", "AssetTag", "DeviceType", "Manufacturer", "Model",
        "PurchaseDate", "WarrantyEnd", "AssignedTo", "AssignmentDate",
        "Location", "Condition", "ReplacementEligible",
        "NumberofReplacement", "AssignedFlag", "WarrantyStatus",
        "WarrantyCovered", "WarrantyDaysRemaining",
        "WarrantyExpiring90Days", "ActionRequired", "Cost", "Notes",
        "IsActive", "UserType", "Status"
    ]

    if request.method == "POST":
        field_values = {}
        for field in editable_fields:
            field_values[field] = request.form.get(field, "").strip()

        if not field_values["DeviceID"]:
            return page(
                "Add Device",
                """
                <div class="header"><h1>💻 Add New Device</h1></div>
                <div class="content">
                    <a class="back" href="/devices">← Devices</a>
                    <p style="color:red;">You must provide a Device ID to add a new record.</p>
                    <a class="btn" href="/devices/add">Back to add form</a>
                </div>
                """
            )

        columns = ", ".join(f'"{field}"' for field in editable_fields)
        values_placeholder = ", ".join(["%s"] * len(editable_fields))
        values = [field_values[field] for field in editable_fields]

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                f'INSERT INTO public.devices ({columns}) VALUES ({values_placeholder})',
                values
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()

        return redirect("/devices")

    form_fields = ""
    for field in editable_fields:
        form_fields += f"""
        <p><b>{field}</b></p>
        <input type=\"text\" name=\"{field}\" style=\"width:400px;padding:8px;\"> \
        <br><br>
        """

    return page(
        "Add Device",
        f"""
        <div class="header"><h1>💻 Add New Device</h1></div>
        <div class="content">
            <a class="back" href="/devices">← Devices</a>
            <form method="POST">
                {form_fields}
                <button type="submit">Add Device</button>
            </form>
        </div>
        """
    )


@app.route("/hotspots/edit/<hotspot_id>", methods=["GET", "POST"])
@login_required
def edit_hotspot(hotspot_id):
    # Edit an existing hotspot movement record using a shared edit renderer.
    editable_fields = ["DeviceID", "FromUser", "ToUser", "Date", "Reason"]

    if request.method == "POST":
        save_edit_fields("hotspots", "MovementID", hotspot_id, editable_fields)
        return redirect("/hotspots")

    hotspot = fetch_single_record("hotspots", "MovementID", hotspot_id)
    if not hotspot:
        return "Hotspot not found."

    return render_edit_page(
        "Edit Hotspot",
        "📶",
        "/hotspots",
        hotspot.get("MovementID", hotspot_id),
        hotspot,
        editable_fields
    )


@app.route("/hotspots/edit", methods=["GET", "POST"])
@login_required
def edit_hotspot_lookup():
    # Lookup page for hotspot edits by MovementID.
    # Includes a quick path to add a new hotspot record instead.
    if request.method == "POST":
        hotspot_id = request.form.get("hotspot_id", "").strip()

        if not hotspot_id:
            return page(
                "Edit Hotspot",
                """
                <div class="header"><h1>📶 Edit Hotspot</h1></div>
                <div class="content">
                    <a class="back" href="/hotspots">← Hotspots</a>
                    <p>Please enter a Movement ID to edit.</p>
                    <form method="POST">
                        <p><b>Movement ID</b></p>
                        <input type="text" name="hotspot_id" style="width:400px;padding:8px;">
                        <br><br>
                        <button type="submit">Open Hotspot</button>
                    </form>
                    <br>
                    <a class="btn" href="/hotspots/add">Add New Hotspot</a>
                </div>
                """
            )

        return redirect(f"/hotspots/edit/{hotspot_id}")

    return page(
        "Edit Hotspot",
        """
        <div class="header">
            <h1>📶 Edit Hotspot</h1>
        </div>

        <div class="content">
            <a class="back" href="/hotspots">← Hotspots</a>

            <p>Enter a Movement ID to open the hotspot editor.</p>

            <form method="POST">
                <p><b>Movement ID</b></p>
                <input type="text" name="hotspot_id" style="width:400px;padding:8px;">
                <br><br>
                <button type="submit">Open Hotspot</button>
            </form>
            <br>
            <a class="btn" href="/hotspots/add">Add New Hotspot</a>
        </div>
        """
    )


@app.route("/staff/edit/<staff_id>", methods=["GET", "POST"])
@login_required
def edit_staff(staff_id):
    editable_fields = ["FullName", "Department", "Email"]

    if request.method == "POST":
        save_edit_fields("staff", "StaffID", staff_id, editable_fields)
        return redirect("/staff")

    staff = fetch_single_record("staff", "StaffID", staff_id)
    if not staff:
        return "Staff member not found."

    return render_edit_page(
        "Edit Staff",
        "👥",
        "/staff",
        staff.get("FullName", staff_id),
        staff,
        editable_fields
    )


@app.route("/students/edit/<student_id>", methods=["GET", "POST"])
@login_required
def edit_student(student_id):
    editable_fields = ["FullName", "GradeLevel", "EnrollmentStatus", "ParentEmail"]

    if request.method == "POST":
        save_edit_fields("students", "StudentID", student_id, editable_fields)
        return redirect("/students")

    student = fetch_single_record("students", "StudentID", student_id)
    if not student:
        return "Student not found."

    return render_edit_page(
        "Edit Student",
        "🎓",
        "/students",
        student.get("FullName", student_id),
        student,
        editable_fields
    )


@app.route("/staff/edit", methods=["GET", "POST"])
@login_required
def edit_staff_lookup():
    # Lookup page for staff edits by StaffID.
    # Users may also follow the add-staff link from this page.
    if request.method == "POST":
        staff_id = request.form.get("staff_id", "").strip()
        if not staff_id:
            return page(
                "Edit Staff",
                """
                <div class="header"><h1>👥 Edit Staff</h1></div>
                <div class="content">
                    <a class="back" href="/staff">← Staff</a>
                    <p>Please enter a Staff ID to edit.</p>
                    <form method="POST">
                        <p><b>Staff ID</b></p>
                        <input type="text" name="staff_id" style="width:400px;padding:8px;">
                        <br><br>
                        <button type="submit">Open Staff</button>
                    </form>
                    <br>
                    <a class="btn" href="/staff/add">Add New Staff</a>
                </div>
                """
            )

        return redirect(f"/staff/edit/{staff_id}")

    return page(
        "Edit Staff",
        """
        <div class="header"><h1>👥 Edit Staff</h1></div>
        <div class="content">
            <a class="back" href="/staff">← Staff</a>
            <form method="POST">
                <p><b>Staff ID</b></p>
                <input type="text" name="staff_id" style="width:400px;padding:8px;">
                <br><br>
                <button type="submit">Open Staff</button>
            </form>
            <br>
            <a class="btn" href="/staff/add">Add New Staff</a>
        </div>
        """
    )


@app.route("/students/edit", methods=["GET", "POST"])
@login_required
def edit_student_lookup():
    # Display a lookup form where users can enter a StudentID to edit.
    # Also provides a button to add a new student if needed.
    if request.method == "POST":
        student_id = request.form.get("student_id", "").strip()
        if not student_id:
            return page(
                "Edit Student",
                """
                <div class="header"><h1>🎓 Edit Student</h1></div>
                <div class="content">
                    <a class="back" href="/students">← Students</a>
                    <p>Please enter a Student ID to edit.</p>
                    <form method="POST">
                        <p><b>Student ID</b></p>
                        <input type="text" name="student_id" style="width:400px;padding:8px;">
                        <br><br>
                        <button type="submit">Open Student</button>
                    </form>
                    <br>
                    <a class="btn" href="/students/add">Add New Student</a>
                </div>
                """
            )

        return redirect(f"/students/edit/{student_id}")

    return page(
        "Edit Student",
        """
        <div class="header"><h1>🎓 Edit Student</h1></div>
        <div class="content">
            <a class="back" href="/students">← Students</a>
            <form method="POST">
                <p><b>Student ID</b></p>
                <input type="text" name="student_id" style="width:400px;padding:8px;">
                <br><br>
                <button type="submit">Open Student</button>
            </form>
            <br>
            <a class="btn" href="/students/add">Add New Student</a>
        </div>
        """
    )


@app.route("/staff/add", methods=["GET", "POST"])
@login_required
def add_staff():
    # Display a form for creating a new staff record.
    # Uses a fixed field set and inserts a new row in the staff table.
    fields = ["StaffID", "FullName", "Department", "Email"]

    if request.method == "POST":
        values = [request.form.get(field, "").strip() for field in fields]

        if not values[0]:
            return page(
                "Add Staff",
                """
                <div class="header"><h1>👥 Add New Staff</h1></div>
                <div class="content">
                    <a class="back" href="/staff">← Staff</a>
                    <p style="color:red;">Staff ID is required.</p>
                    <a class="btn" href="/staff/add">Back to Add Staff</a>
                </div>
                """
            )

        columns = ", ".join([f'"{field}"' for field in fields])
        values_placeholder = ", ".join(["%s"] * len(fields))

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                f'INSERT INTO public.staff ({columns}) VALUES ({values_placeholder})',
                values
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()

        return redirect("/staff")

    form_inputs = ""
    for field in fields:
        form_inputs += f"""
            <p><b>{field}</b></p>
            <input type=\"text\" name=\"{field}\" style=\"width:400px;padding:8px;\">
            <br><br>
        """

    return page(
        "Add Staff",
        f"""
        <div class="header"><h1>👥 Add New Staff</h1></div>
        <div class="content">
            <a class="back" href="/staff">← Staff</a>
            <form method="POST">
                {form_inputs}
                <button type="submit">Add Staff</button>
            </form>
        </div>
        """
    )


@app.route("/students/add", methods=["GET", "POST"])
@login_required
def add_student():
    # Display a form for creating a new student record.
    # Adds the row to the students table on form submission.
    fields = ["StudentID", "FullName", "GradeLevel", "EnrollmentStatus", "ParentEmail"]

    if request.method == "POST":
        values = [request.form.get(field, "").strip() for field in fields]

        if not values[0]:
            return page(
                "Add Student",
                """
                <div class="header"><h1>🎓 Add New Student</h1></div>
                <div class="content">
                    <a class="back" href="/students">← Students</a>
                    <p style="color:red;">Student ID is required.</p>
                    <a class="btn" href="/students/add">Back to Add Student</a>
                </div>
                """
            )

        columns = ", ".join([f'"{field}"' for field in fields])
        values_placeholder = ", ".join(["%s"] * len(fields))

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                f'INSERT INTO public.students ({columns}) VALUES ({values_placeholder})',
                values
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()

        return redirect("/students")

    form_inputs = ""
    for field in fields:
        form_inputs += f"""
            <p><b>{field}</b></p>
            <input type=\"text\" name=\"{field}\" style=\"width:400px;padding:8px;\">
            <br><br>
        """

    return page(
        "Add Student",
        f"""
        <div class="header"><h1>🎓 Add New Student</h1></div>
        <div class="content">
            <a class="back" href="/students">← Students</a>
            <form method="POST">
                {form_inputs}
                <button type="submit">Add Student</button>
            </form>
        </div>
        """
    )


@app.route("/hotspots/add", methods=["GET", "POST"])
@login_required
def add_hotspot():
    # Display a form for adding a new hotspot movement record.
    fields = ["MovementID", "DeviceID", "FromUser", "ToUser", "Date", "Reason"]

    if request.method == "POST":
        values = [request.form.get(field, "").strip() for field in fields]

        if not values[0]:
            return page(
                "Add Hotspot",
                """
                <div class="header"><h1>📶 Add New Hotspot</h1></div>
                <div class="content">
                    <a class="back" href="/hotspots">← Hotspots</a>
                    <p style="color:red;">Movement ID is required.</p>
                    <a class="btn" href="/hotspots/add">Back to Add Hotspot</a>
                </div>
                """
            )

        columns = ", ".join([f'"{field}"' for field in fields])
        values_placeholder = ", ".join(["%s"] * len(fields))

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                f'INSERT INTO public.hotspots ({columns}) VALUES ({values_placeholder})',
                values
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()

        return redirect("/hotspots")

    form_inputs = ""
    for field in fields:
        form_inputs += f"""
            <p><b>{field}</b></p>
            <input type=\"text\" name=\"{field}\" style=\"width:400px;padding:8px;\">
            <br><br>
        """

    return page(
        "Add Hotspot",
        f"""
        <div class="header"><h1>📶 Add New Hotspot</h1></div>
        <div class="content">
            <a class="back" href="/hotspots">← Hotspots</a>
            <form method="POST">
                {form_inputs}
                <button type="submit">Add Hotspot</button>
            </form>
        </div>
        """
    )


@app.route("/bulk-upload", methods=["GET", "POST"])
@login_required
def bulk_upload():
    # Upload and bulk import spreadsheet data for supported tables.
    # Admin-only page that accepts CSV files and merges rows.

    if current_user.role != "Admin":
        return "Access denied. Admins only."

    table_configs = {
        "devices": {
            "pk": "AssetTag",
            "columns": ["DeviceID", "AssetTag", "DeviceType", "Manufacturer", "Model", "PurchaseDate", "WarrantyEnd", "AssignedTo", "AssignmentDate", "Location", "Condition", "ReplacementEligible", "NumberofReplacement", "AssignedFlag", "WarrantyStatus", "WarrantyCovered", "WarrantyDaysRemaining", "WarrantyExpiring90Days", "ActionRequired", "Cost", "Notes", "IsActive", "UserType", "Status"]
        },
        "staff": {
            "pk": "StaffID",
            "columns": ["StaffID", "FullName", "Department", "Email"]
        },
        "students": {
            "pk": "StudentID",
            "columns": ["StudentID", "FullName", "GradeLevel", "EnrollmentStatus", "ParentEmail"]
        },
        "movements": {
            "pk": "MovementID",
            "columns": ["MovementID", "AssetTag", "MovementDate", "MovementType", "FromLocation", "FromAssignedTo", "ToAssignedTo", "LatestFlag", "KEY", "Note"]
        }
    }

    message = ""

    if request.method == "POST":
        table_name = request.form.get("table")
        uploaded_file = request.files.get("file")

        # Validate the selected target table and file upload.
        if table_name not in table_configs:
            message = "Invalid table selected."
        elif not uploaded_file:
            message = "No file selected."
        elif not is_allowed_upload_file(uploaded_file):
            message = "Only .csv uploads are allowed."
        else:
            uploaded_file.stream.seek(0)
            headers, rows, error = read_upload_rows(uploaded_file)

            if error:
                message = error
            else:
                config = table_configs[table_name]
                pk = config["pk"]
                allowed_columns = config["columns"]

                valid_headers = [h for h in headers if h in allowed_columns]

                if pk not in valid_headers:
                    message = f"Upload failed. Your CSV file must include the ID column: {pk}"
                else:
                    conn = get_db_connection()
                    cur = conn.cursor()

                inserted = 0
                updated = 0
                skipped = 0
                seen_keys = set()

                for row in rows:
                    row_data = dict(zip(headers, row))

                    clean_data = {}

                    for col in valid_headers:
                        value = row_data.get(col)

                        if value is not None and value != "":
                            clean_data[col] = value

                    if not clean_data:
                        skipped += 1
                        continue

                    pk_value = clean_data.get(pk)
                    if table_name == "devices":
                        asset_tag_check = clean_data.get("AssetTag")

                        if asset_tag_check in seen_keys:
                             skipped += 1
                             continue

                        seen_keys.add(asset_tag_check)

                    if pk_value:
                        update_cols = [c for c in clean_data.keys() if c != pk]

                        if update_cols:
                            set_clause = ", ".join([f'"{c}" = %s' for c in update_cols])
                            values = [clean_data[c] for c in update_cols]
                            values.append(pk_value)

                            cur.execute(
                                f'UPDATE public.{table_name} SET {set_clause} WHERE "{pk}" = %s',
                                values
                            )

                            if cur.rowcount > 0:
                                updated += 1
                            else:
                                cols = list(clean_data.keys())
                                col_names = ", ".join([f'"{c}"' for c in cols])
                                placeholders = ", ".join(["%s"] * len(cols))
                                insert_values = [clean_data[c] for c in cols]

                                cur.execute(
                                    f'INSERT INTO public.{table_name} ({col_names}) VALUES ({placeholders})',
                                    insert_values
                                )

                                inserted += 1
                        else:
                            skipped += 1

                    else:
                        insert_data = {k: v for k, v in clean_data.items() if k != pk}

                        if insert_data:
                            cols = list(insert_data.keys())
                            col_names = ", ".join([f'"{c}"' for c in cols])
                            placeholders = ", ".join(["%s"] * len(cols))
                            insert_values = [insert_data[c] for c in cols]

                            cur.execute(
                                f'INSERT INTO public.{table_name} ({col_names}) VALUES ({placeholders})',
                                insert_values
                            )

                            inserted += 1
                        else:
                            skipped += 1

                conn.commit()
                cur.close()
                conn.close()

                message = f"Upload complete. Inserted: {inserted}, Updated: {updated}, Skipped: {skipped}"

    return f"""
    <div class="header">
        <h1>⬆️ Bulk Upload</h1>
    </div>

    <div class="content">

        <a class="back" href="/">← Dashboard</a>

        <h2>Upload CSV Files</h2>

        <form method="POST" enctype="multipart/form-data">

            <p>Table</p>

           <select name="table">
    <option value="devices">Devices</option>
    <option value="staff">Staff</option>
    <option value="students">Students</option>
    <option value="movements">Movements</option>
    <option value="hotspots">Hotspots</option>
</select>

            <br><br>

            <p>Supported format: .csv only.</p>
            <input type="file" name="file" accept=".csv">

            <br><br>

            <button type="submit">Upload</button>

        </form>

        <p style="color:green;font-weight:bold;">
            {message}
        </p>

    </div>
    """
@app.route("/movements")
@login_required
def movements():
    # Display the movement history table with search and navigation controls.
    return make_table_page(
        "Movement History",
        "🔄",
        "movements",
        [
            "MovementID",
            "AssetTag",
            "MovementDate",
            "MovementType",
            "Fromlocation",
            "FromAssignedTo",
            "ToAssignedTo",
            "LatestFlag",
            "KEY",
            "Note"
        ]
    )
@app.route("/movement-entry", methods=["GET", "POST"])
@login_required
def movement_entry():
    # Process device movements such as Assign, Return, Transfer, Repair, or Lost.
    # Updates both the movements history and device status in a single request.

    message = ""

    if request.method == "POST":

        asset_tag = request.form.get("asset_tag", "").strip()
        movement_type = request.form.get("movement_type", "").strip()
        staff_assigned = request.form.get("staff_assigned", "").strip()
        student_assigned = request.form.get("student_assigned", "").strip()

        assigned_to = staff_assigned if staff_assigned else student_assigned
        user_type = "Staff" if staff_assigned else "Student"

        location = request.form.get("location", "").strip()
        note = request.form.get("note", "").strip()

        if movement_type == "Return":
            assigned_to = ""
            user_type = ""
            location = "Warehouse"
            device_status = "Unassigned"
        elif movement_type == "Assign":
            device_status = "Assigned"
        elif movement_type == "Transfer":
            device_status = "Assigned"
        elif movement_type == "Repair":
            device_status = "Repair"
        elif movement_type == "Lost":
            device_status = "Lost"
        else:
            device_status = movement_type

        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cur.execute("""
            SELECT "AssignedTo", "Location"
            FROM public.devices
            WHERE "AssetTag" = %s
        """, (asset_tag,))

        device = cur.fetchone()

        if not device:
            message = "Device not found. Check the Asset Tag."
        else:
            from_assigned_to = device.get("AssignedTo", "")
            from_location = device.get("Location", "")

            cur.execute("""
                INSERT INTO public.movements
                (
                    "AssetTag",
                    "MovementDate",
                    "MovementType",
                    "Fromlocation",
                    "FromAssignedTo",
                    "ToAssignedTo",
                    "LatestFlag",
                    "KEY",
                    "Note"
                )
                VALUES
                (
                    %s,
                    TO_CHAR(CURRENT_DATE, 'MM/DD/YYYY'),
                    %s,
                    %s,
                    %s,
                    %s,
                    'Latest',
                    %s,
                    %s
                )
            """, (
                asset_tag,
                movement_type,
                from_location,
                from_assigned_to,
                assigned_to,
                asset_tag + "-Latest",
                note
            ))

            cur.execute("""
                UPDATE public.devices
                SET
                    "AssignedTo" = %s,
                    "Location" = %s,
                    "AssignmentDate" = TO_CHAR(CURRENT_DATE, 'MM/DD/YYYY'),
                    "Status" = %s,
                    "UserType" = %s
                WHERE "AssetTag" = %s
            """, (
                assigned_to,
                location,
                device_status,
                user_type,
                asset_tag
            ))

            conn.commit()
            message = "Movement processed successfully."

        cur.close()
        conn.close()

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        SELECT "StaffID", "FullName"
        FROM public.staff
        ORDER BY "FullName"
    """)
    staff_rows = cur.fetchall()

    cur.execute("""
        SELECT "StudentID", "FullName"
        FROM public.students
        ORDER BY "FullName"
    """)
    student_rows = cur.fetchall()

    cur.close()
    conn.close()

    staff_options = ""
    for staff in staff_rows:
        staff_options += f'<option value="{staff["StaffID"]}">{staff["FullName"]}</option>'

    student_options = ""
    for student in student_rows:
        student_options += f'<option value="{student["StudentID"]}">{student["FullName"]}</option>'

    return f"""
    <div class="header">
        <h1>📦 Movement Entry</h1>
        <p>Assign, Return, Transfer, Repair, Lost</p>
    </div>

    <div class="content">

        <a class="back" href="/">← Dashboard</a>

        <form method="POST">

            <p><b>Asset Tag</b></p>
            <input type="text" name="asset_tag" style="width:400px;padding:8px;" required>

            <p><b>Movement Type</b></p>
            <select name="movement_type" style="width:400px;padding:8px;">
                <option value="Assign">Assign</option>
                <option value="Return">Return</option>
                <option value="Transfer">Transfer</option>
                <option value="Repair">Repair</option>
                <option value="Lost">Lost</option>
            </select>

            <p><b>Assign to Staff</b></p>
            <select name="staff_assigned" style="width:400px;padding:8px;">
                <option value="">-- Select Staff Member --</option>
                {staff_options}
            </select>

            <p><b>Assign to Student</b></p>
            <select name="student_assigned" style="width:400px;padding:8px;">
                <option value="">-- Select Student --</option>
                {student_options}
            </select>

            <p><b>Location</b></p>
            <select name="location" style="width:400px;padding:8px;">
                <option value="Warehouse">Warehouse</option>
                <option value="District Office">District Office</option>
                <option value="High School">High School</option>
                <option value="Middle School">Middle School</option>
                <option value="Elementary School">Elementary School</option>
                <option value="Repair Vendor">Repair Vendor</option>
            </select>

            <p><b>Notes</b></p>
            <textarea name="note" style="width:500px;height:100px;"></textarea>

            <br><br>

            <button class="btn" type="submit">
                Process Movement
            </button>

        </form>

        <br>

        <div style="color:green;font-weight:bold;">
            {message}
        </div>

    </div>
    """

@app.route("/devices")
@login_required
def devices():
    # Show the device inventory table with the generic table renderer.
    # The action toolbar includes an add/edit lookup for devices.
    return make_table_page(
        "Device Inventory",
        "💻",
        "devices",
        [
            "DeviceID", "AssetTag", "DeviceType", "Manufacturer", "Model",
            "PurchaseDate", "WarrantyEnd", "AssignedTo", "AssignmentDate",
            "Location", "Condition", "ReplacementEligible",
            "NumberofReplacement", "AssignedFlag", "WarrantyStatus",
            "WarrantyCovered", "WarrantyDaysRemaining",
            "WarrantyExpiring90Days", "ActionRequired", "Cost", "Notes",
            "IsActive", "UserType", "Status"
        ]
    )



@app.route("/staff")
@login_required
def staff():
    # Show the staff directory table and provide the edit/add lookup action.
    return make_table_page(
        "Staff",
        "👥",
        "staff",
        ["StaffID", "FullName", "Email", "Department"]
    )

@app.route("/students")
@login_required
def students():
    # Show the student directory table and provide the edit/add lookup action.
    return make_table_page(
        "Students",
        "🎓",
        "students",
        ["StudentID", "FullName", "GradeLevel", "EnrollmentStatus", "ParentEmail"]
    )

@app.route("/hotspots")
@login_required
def hotspots():
    # Show the hotspot movements table and allow navigation to hotspot add/edit flows.
    return make_table_page(
        "Hotspots",
        "📶",
        "hotspots",
        [
            "MovementID",
            "DeviceID",
            "FromUser",
            "ToUser",
            "Date",
            "Reason"
        ]
    )

@app.route("/depreciation")
@login_required
def depreciation():
    # Depreciation is restricted — block access for non-admins (e.g., engineers).
    admin_check = require_admin()
    if admin_check:
        return admin_check

    return make_table_page(
        "Depreciation",
        "💰",
        "depreciation",
        [
            "Cost",
            "UsefulLifeYears",
            "SalvageValue",
            "DepreciationMethod",
            "AnnualDepreciation",
            "AccumulatedDepreciation",
            "NetBookValue",
            "deviceId",
            "DeviceStatus"
        ]
    )
    
@app.route("/repairs")
@login_required
def repairs():
    return make_table_page(
        "Repair Tracking",
        "",
        "repairs",
        [
            "TicketID",
            "AssetTag",
            "DateReported",
            "IssueType",
            "RepairStatus",
            "Vendor",
            "DateResolved",
            "Cost",
            "DeviceCondition",
            "ProductID",
            "Priority",
            "Action",
            "RepairCategory",
            "WarrantyStatus",
            "DeviceAge"
        ]
    )


@app.route("/repairs/add", methods=["GET", "POST"])
@login_required
def add_repair():
    repair_fields = [
        "TicketID",
        "AssetTag",
        "DateReported",
        "IssueType",
        "RepairStatus",
        "Vendor",
        "DateResolved",
        "Cost",
        "DeviceCondition",
        "ProductID",
        "Priority",
        "Action",
        "RepairCategory",
        "WarrantyStatus",
        "DeviceAge"
    ]

    if request.method == "POST":
        values = []

        for field in repair_fields:
            value = request.form.get(field, "").strip()
            values.append(value if value else None)

        if values[0] is None:
            return page(
                "Add Repair Error",
                """
                <div class="header">
                    <h1>Add Repair Error</h1>
                </div>

                <div class="content">
                    <div class="card">
                        <p>TicketID is required.</p>
                        <a class="btn" href="/repairs/add">
                            Return to Add Repair
                        </a>
                    </div>
                </div>
                """
            )

        try:
            values[0] = int(values[0])

            if values[7] is not None:
                values[7] = int(values[7])

        except ValueError:
            return page(
                "Add Repair Error",
                """
                <div class="header">
                    <h1>Add Repair Error</h1>
                </div>

                <div class="content">
                    <div class="card">
                        <p>TicketID and Cost must contain numbers only.</p>
                        <a class="btn" href="/repairs/add">
                            Return to Add Repair
                        </a>
                    </div>
                </div>
                """
            )

        conn = get_db_connection()
        cur = conn.cursor()

        try:
            column_sql = ", ".join(
                f'"{field}"' for field in repair_fields
            )

            placeholders = ", ".join(
                ["%s"] * len(repair_fields)
            )

            cur.execute(
                f'''
                INSERT INTO public."repairs" ({column_sql})
                VALUES ({placeholders})
                ''',
                values
            )

            conn.commit()

        except Exception as error:
            conn.rollback()
            error_message = escape(str(error))

            return page(
                "Add Repair Error",
                f"""
                <div class="header">
                    <h1>Add Repair Error</h1>
                </div>

                <div class="content">
                    <div class="card">
                        <p>{error_message}</p>
                        <a class="btn" href="/repairs/add">
                            Return to Add Repair
                        </a>
                    </div>
                </div>
                """
            )

        finally:
            cur.close()
            conn.close()

        return redirect("/repairs")

    blank_repair = {
        field: ""
        for field in repair_fields
    }

    return render_edit_page(
        "Add Repair",
        "",
        "/repairs",
        "Add Repair",
        blank_repair,
        repair_fields
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)